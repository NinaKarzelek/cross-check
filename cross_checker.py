# This code loads an Excel workbook (Detail Item Schedule.xlsx), finds the cell with the value 'EGS_Name', and then looks for all non-empty cells below it.
# For each non-empty cell, it searches all Excel workbooks in a specified directory (workbooks/) for matching cells and highlights them in green.
# It then generates a report for each workbook in the specified directory, showing the total number of cells and the number of cells that were highlighted in green.
# This code calculate all unempty cells for totals in report

import os

import openpyxl

# Load the first workbook and select the first sheet
#IMPORTANT - update the file path 
workbook1 = openpyxl.load_workbook('C:\\Users\\Nina.Karzelek\\Desktop\\python\\cross-check\\PA - M - Detail Item Schedule.xlsx')
sheet1 = workbook1.active

# Find the cell with the value 'EGS_Name'
egs_name_cell = None
for row in sheet1.iter_rows():
    for cell in row:
        if cell.value == 'EGS_Name':
            egs_name_cell = cell
            break
    if egs_name_cell:
        break

# If the cell is found, highlight matching cells in all workbooks in the specified directory
if egs_name_cell:
    egs_name_col = egs_name_cell.column_letter
    for cell in sheet1[egs_name_col]:
        if cell.row <= egs_name_cell.row or not cell.value:
            continue  # Skip cells above the 'EGS_Name' cell or empty cells

        # Find matching cells in all workbooks in the specified directory
        #IMPORTANT - update the file path 
        directory = 'C:\\Users\\Nina.Karzelek\\Desktop\\python\\cross-check'
        for filename in os.listdir(directory):
            if not filename.endswith('.xlsx'):
                continue
            filepath = os.path.join(directory, filename)
            workbook2 = openpyxl.load_workbook(filepath)
            sheet2 = workbook2.active
            for row2 in sheet2.iter_rows():
                for cell2 in row2:
                    if cell2.value == cell.value:
                        cell2.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            workbook2.save(filepath)
else:
    print("Cell with header 'EGS_Name' not found.")

# Generate a single report for all workbooks in the specified directory
#IMPORTANT - update the file path 
directory = 'C:\\Users\\Nina.Karzelek\\Desktop\\python\\cross-check'
total_cells_dict = {}  # dictionary to store the total cells count for each workbook
green_cells_dict = {}  # dictionary to store the green cells count for each workbook
for filename in os.listdir(directory):
    if not filename.endswith('.xlsx'):
        continue
    filepath = os.path.join(directory, filename)
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    total_cells = 0
    green_cells = 0
    egs_name_cell = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'EGS_Name':
                egs_name_cell = cell
                break
        if egs_name_cell:
            break
    if egs_name_cell:
        egs_name_col = egs_name_cell.column_letter
        for cell in sheet[egs_name_col]:
            if cell.row <= egs_name_cell.row or not cell.value:
                continue  # Skip cells above the 'EGS_Name' cell or empty cells
            total_cells += 1
            if cell.fill.start_color.index == openpyxl.styles.colors.Color(rgb='00FF00').rgb:
                green_cells += 1
    total_cells_dict[filename] = total_cells
    green_cells_dict[filename] = green_cells

import datetime
# create a timestamp
now = datetime.datetime.now()
timestamp = now.strftime("%Y-%m-%d")

report_workbook = openpyxl.Workbook()
report_sheet = report_workbook.active
report_sheet['A1'] = 'File Name'
report_sheet['B1'] = 'Total Number of  Model Elements'
report_sheet['C1'] = 'Matched Elements'
report_sheet['D1'] = 'Percentage of Model Elements which have a match in Schematic'

row_num = 2
for filename, total_cells in total_cells_dict.items():
    report_sheet.cell(row=row_num, column=1, value=filename)
    if filename != 'Detail Item Schedule.xlsx':  # skip the first workbook
        report_sheet.cell(row=row_num, column=2, value=total_cells)
        green_cells = green_cells_dict[filename]
        report_sheet.cell(row=row_num, column=3, value=green_cells)
        percentage = round((green_cells / total_cells), 2)
        report_sheet.cell(row=row_num, column=4, value=f'{percentage}')
    row_num += 1

#IMPORTANT - update the file path 
report_workbook.save(f'C:\\Users\\Nina.Karzelek\\Desktop\\python\\cross-check\\report_{timestamp}.xlsx')
